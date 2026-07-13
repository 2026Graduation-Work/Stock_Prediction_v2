"""데이터 수집기 (무료 소스 우선, 실패/키 없음 시 샘플 폴백).

뉴스   : 네이버 검색 OpenAPI(키 있으면) → HTML 크롤(키 없을 때) → 샘플
소셜   : 네이버 금융 종목토론 크롤링          (키 불필요, best-effort)
시세   : FinanceDataReader                   (키 불필요)
재무제표: OpenDartReader / Open DART API      (무료 키 필요)
어느 단계든 실패하면 sample_data/ 의 동봉 데이터로 대체되어 파이프라인이 끝까지 돈다.
"""
from __future__ import annotations

import datetime as dt
import html
import json
import re
import unicodedata
import warnings
from email.utils import parsedate_to_datetime
from pathlib import Path
from typing import Any

from .config import SETTINGS

SAMPLE_DIR = Path(__file__).parent / "sample_data"
DATA_DIR = Path(__file__).resolve().parents[4] / "data"


def _load_sample(name: str):
    p = SAMPLE_DIR / name
    if p.exists():
        return json.loads(p.read_text(encoding="utf-8"))
    return None


# ── 공통: 날짜 필터 ──────────────────────────────────────────────
def _filter_by_date(items: list[dict], date: str) -> list[dict]:
    """게시일이 해당 날짜(YYYY-MM-DD)인 항목만 남긴다 (point-in-time)."""
    return [it for it in items if str(it.get("date", "")).startswith(date)]


# ── 뉴스 ────────────────────────────────────────────────────────
def collect_news(ticker: str, company_name: str, date: str) -> tuple[list[dict], str]:
    """해당 날짜 '하루치' 상위 10건 기사 리스트와 출처 반환.
    각 항목: {title, summary, url, press, date}.
    우선순위: 로컬 엑셀(data/) → 네이버 검색 OpenAPI(키 있으면) → HTML 크롤 → 샘플."""
    query = company_name or ticker
    try:
        items = _load_excel_news(query, date)
        if items:
            return items, "excel"
    except Exception:
        pass
    if SETTINGS.has_naver:
        try:
            items = _fetch_naver_news_api(query, date)
            if items:
                return items, "naver_api"
        except Exception:
            pass
    try:
        items = _crawl_naver_news(query, date)
        if items:
            return items, "naver"
    except Exception:
        pass
    sample = _load_sample(f"{ticker}_news.json") or _load_sample("default_news.json") or []
    return _filter_by_date(sample, date)[:10], "sample"


def _parse_news_workbook_name(path: Path) -> tuple[str, str, str] | None:
    match = re.fullmatch(r"(.+)_(\d{8})-(\d{8})\.xlsx", path.name)
    if not match:
        return None
    return match.group(1), match.group(2), match.group(3)


def _normalize_yyyymmdd(value: Any) -> str:
    if isinstance(value, dt.datetime | dt.date):
        return value.strftime("%Y%m%d")
    text = str(value or "").strip()
    if not text:
        return ""
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    digits = re.sub(r"\D", "", text)
    return digits[:8]


def _format_iso_date(yyyymmdd: str) -> str:
    return f"{yyyymmdd[:4]}-{yyyymmdd[4:6]}-{yyyymmdd[6:8]}"


def _find_excel_news_file(company_name: str, date: str) -> Path | None:
    target_day = date.replace("-", "")
    normalized_company = unicodedata.normalize("NFC", company_name)
    for path in sorted(DATA_DIR.glob("*.xlsx")):
        parsed = _parse_news_workbook_name(path)
        if parsed is None:
            continue
        file_company, start, end = parsed
        file_company = unicodedata.normalize("NFC", file_company)
        if file_company == normalized_company and start <= target_day <= end:
            return path
    return None


def _load_excel_news(company_name: str, date: str, limit: int = 10) -> list[dict]:
    """data/<회사명>_YYYYMMDD-YYYYMMDD.xlsx에서 해당 날짜 제목을 읽는다."""
    path = _find_excel_news_file(company_name, date)
    if path is None:
        return []

    from openpyxl import load_workbook

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Workbook contains no default style.*",
            category=UserWarning,
        )
        wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if hasattr(ws, "reset_dimensions"):
        ws.reset_dimensions()
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return []

    columns = {str(name).strip(): idx for idx, name in enumerate(header) if name is not None}
    if "일자" not in columns or "제목" not in columns:
        raise RuntimeError(f"엑셀 뉴스 파일 필수 컬럼 누락: {path}")

    date_idx = columns["일자"]
    title_idx = columns["제목"]
    press_idx = columns.get("언론사")
    target_day = date.replace("-", "")
    items: list[dict] = []
    for row in rows:
        row_day = _normalize_yyyymmdd(row[date_idx] if date_idx < len(row) else "")
        if row_day != target_day:
            continue
        title = str(row[title_idx] if title_idx < len(row) else "").strip()
        if not title:
            continue
        press = ""
        if press_idx is not None and press_idx < len(row) and row[press_idx] is not None:
            press = str(row[press_idx]).strip()
        items.append({
            "title": title,
            "summary": "",
            "url": "",
            "press": press,
            "date": _format_iso_date(row_day),
        })
        if len(items) >= limit:
            break
    return items


def _strip_tags(s: str) -> str:
    """네이버 API 제목/요약의 <b> 태그·HTML 엔티티 제거."""
    return html.unescape(re.sub(r"<[^>]+>", "", s or "")).strip()


def _fetch_naver_news_api(query: str, date: str, limit: int = 10) -> list[dict]:
    """네이버 검색 OpenAPI로 해당 날짜 상위 limit건 수집.

    API는 기간 필터를 지원하지 않으므로 sort=date(최신순)로 페이지를 넘기며
    pubDate가 해당 날짜인 기사만 모으고, 해당 날짜보다 과거가 되면 중단한다.
    (주의: API는 최근 1,000건까지만 반환 → 너무 오래된 날짜는 도달 못 할 수 있음.)
    """
    import requests

    headers = {
        "X-Naver-Client-Id": SETTINGS.naver_client_id,
        "X-Naver-Client-Secret": SETTINGS.naver_client_secret,
    }
    items: list[dict] = []
    for start in range(1, 1001, 100):  # 최대 1,000건(100건 × 10페이지)
        r = requests.get(
            "https://openapi.naver.com/v1/search/news.json",
            params={"query": query, "display": 100, "start": start, "sort": "date"},
            headers=headers,
            timeout=10,
        )
        r.raise_for_status()
        arr = r.json().get("items", [])
        if not arr:
            break
        page_days: list[str] = []
        for it in arr:
            try:
                day = parsedate_to_datetime(it["pubDate"]).date().isoformat()
            except Exception:
                continue
            page_days.append(day)
            if day == date:
                items.append({
                    "title": _strip_tags(it.get("title", "")),
                    "summary": _strip_tags(it.get("description", "")),
                    "url": it.get("originallink") or it.get("link", ""),
                    "press": "", "date": date,
                })
                if len(items) >= limit:
                    return items
        if page_days and all(d < date for d in page_days):
            break  # 최신순이라 이 페이지부터 전부 과거 → 해당 날짜 없음
    if not items:
        raise RuntimeError("OpenAPI: 해당 날짜 뉴스 없음")
    return items


def _crawl_naver_news(query: str, date: str, limit: int = 10) -> list[dict]:
    """네이버 뉴스 검색을 해당 날짜 1일로 기간 지정(pd=3, ds=de)해 상위 10건 제목 수집."""
    import requests
    from bs4 import BeautifulSoup

    d_dot = date.replace("-", ".")   # 2026.05.17
    d_num = date.replace("-", "")    # 20260517
    r = requests.get(
        "https://search.naver.com/search.naver",
        params={
            "where": "news", "query": query, "sort": "1",
            "pd": "3", "ds": d_dot, "de": d_dot,          # 기간: 해당 날짜 하루
            "nso": f"so:r,p:from{d_num}to{d_num},a:all",
        },
        headers={"User-Agent": "Mozilla/5.0"},
        timeout=10,
    )
    r.raise_for_status()
    soup = BeautifulSoup(r.text, "html.parser")
    items: list[dict] = []
    for a in soup.select("a.news_tit")[:limit]:
        title = a.get("title") or a.get_text(strip=True)
        if title:
            items.append({"title": title, "summary": "", "url": a.get("href"),
                          "press": "", "date": date})
    if not items:
        raise RuntimeError("해당 날짜 뉴스 없음/파싱 실패 (마크업 변경 가능)")
    return items


# ── 소셜 (네이버 금융 종목토론실) ────────────────────────────────
def collect_social(ticker: str, date: str) -> tuple[list[dict], str]:
    """해당 날짜 '하루치' 게시글 리스트와 출처 반환. 각 항목: {title, date}."""
    try:
        posts = _crawl_naver_board(ticker, date)
        if posts:
            return posts, "naver"
    except Exception:
        pass
    sample = _load_sample(f"{ticker}_social.json") or _load_sample("default_social.json") or []
    return _filter_by_date(sample, date), "sample"


def _crawl_naver_board(ticker: str, date: str, max_pages: int = 50) -> list[dict]:
    """종목토론실을 페이지별로 읽으며 게시일이 해당 날짜인 글만 수집.

    글이 최신순이라, 한 페이지 전체가 해당 날짜보다 과거가 되면 중단한다.
    """
    import requests
    from bs4 import BeautifulSoup

    posts: list[dict] = []
    for page in range(1, max_pages + 1):
        r = requests.get(
            "https://finance.naver.com/item/board.naver",
            params={"code": ticker, "page": page},
            headers={"User-Agent": "Mozilla/5.0"},
            timeout=10,
        )
        r.raise_for_status()
        r.encoding = "euc-kr"  # 네이버 금융은 euc-kr
        soup = BeautifulSoup(r.text, "html.parser")

        page_days: list[str] = []
        has_row = False
        for tr in soup.select("table.type2 tr"):
            a = tr.select_one("td.title a")
            span = tr.select_one("span.tah")  # 날짜 셀 (예: 2026.05.17 14:30)
            if not a or not span:
                continue
            has_row = True
            day = span.get_text(strip=True)[:10].replace(".", "-")  # 2026-05-17
            page_days.append(day)
            if day == date:
                title = a.get("title") or a.get_text(strip=True)
                if title:
                    posts.append({"title": title, "date": date})
        if not has_row:
            break
        if page_days and all(d < date for d in page_days):
            break  # 이 페이지부터 전부 과거 → 더 봐도 해당 날짜 없음
    if not posts:
        raise RuntimeError("해당 날짜 게시글 없음/파싱 실패")
    return posts


# ── 재무제표 + 시세 ─────────────────────────────────────────────
def collect_financials(ticker: str, date: str) -> tuple[dict, str]:
    """표준화 재무 dict와 출처('dart'/'sample') 반환. 시세는 가능하면 FDR로 덮어쓴다."""
    data: dict | None = None
    source = "sample"
    if SETTINGS.has_dart:
        try:
            data = _fetch_dart_financials(ticker)
            source = "dart"
        except Exception:
            data = None
    if not data:
        data = (
            _load_sample(f"{ticker}_financials.json")
            or _load_sample("default_financials.json")
            or {}
        )
        data = dict(data)  # 복사
        source = "sample"

    price = _fetch_price(ticker, date)
    if price is not None:
        data["price"] = price
        data["_price_source"] = "fdr"
    return data, source


def _fetch_price(ticker: str, date: str) -> float | None:
    """date 이전 마지막 종가 (FinanceDataReader, 키 불필요)."""
    try:
        import FinanceDataReader as fdr

        end = dt.date.fromisoformat(date)
        start = end - dt.timedelta(days=SETTINGS.price_lookback_days)
        df = fdr.DataReader(ticker, start.isoformat(), end.isoformat())
        if df is None or len(df) == 0:
            return None
        return float(df["Close"].iloc[-1])
    except Exception:
        return None


# DART 표준계정 → 표준 키 매핑 (best-effort)
_DART_MAP = {
    "매출액": "revenue", "수익(매출액)": "revenue", "영업수익": "revenue",
    "영업이익": "operating_profit", "영업이익(손실)": "operating_profit",
    "당기순이익": "net_income", "당기순이익(손실)": "net_income",
    "자산총계": "total_assets", "부채총계": "total_liabilities", "자본총계": "total_equity",
    "유동자산": "current_assets", "유동부채": "current_liabilities",
    "재고자산": "inventories", "이익잉여금": "retained_earnings",
}


def _fetch_dart_financials(ticker: str) -> dict:
    """OpenDartReader로 최근 사업보고서 재무제표를 표준 dict로 변환.

    DART 계정명/구조가 회사마다 달라 best-effort 매핑이며,
    매핑 실패 항목은 결측(None)으로 남아 지표가 부분 계산된다.
    """
    import OpenDartReader

    dart = OpenDartReader(SETTINGS.dart_api_key)
    year = dt.date.today().year - 1  # 직전 사업연도
    fs = dart.finstate_all(ticker, year)  # 연결재무제표 전체 계정
    if fs is None or len(fs) == 0:
        raise RuntimeError("DART 재무제표 없음")

    out: dict = {}
    for _, row in fs.iterrows():
        name = str(row.get("account_nm", "")).strip()
        key = _DART_MAP.get(name)
        if not key or key in out:
            continue
        raw = str(row.get("thstrm_amount", "")).replace(",", "")
        try:
            out[key] = float(raw)
        except ValueError:
            continue
    if "revenue" not in out:
        raise RuntimeError("DART 매핑 실패")

    # 전년 동기 (성장률용)
    for _, row in fs.iterrows():
        name = str(row.get("account_nm", "")).strip()
        key = _DART_MAP.get(name)
        if key in ("revenue", "operating_profit", "net_income"):
            raw = str(row.get("frmtrm_amount", "")).replace(",", "")
            try:
                out.setdefault(f"{key}_prev", float(raw))
            except ValueError:
                pass

    # 발행주식수
    try:
        sh = dart.report(ticker, "주식총수", year)
        out["shares_outstanding"] = float(str(sh.iloc[0]["istc_totqy"]).replace(",", ""))
    except Exception:
        pass

    # 업종 평균 PER/PBR은 별도 소스 필요 → 샘플 기본값 유지
    sample = _load_sample(f"{ticker}_financials.json") or {}
    out.setdefault("sector_per", sample.get("sector_per"))
    out.setdefault("sector_pbr", sample.get("sector_pbr"))
    return out
